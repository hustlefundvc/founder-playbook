#!/usr/bin/env python3
"""
Build the Mode 1 own-network connector spreadsheet from scraped LinkedIn results.

Why this exists: every Mode 1 run does the same pivot (target -> mutuals becomes
connector -> targets) and the same workbook layout. Doing it once here keeps the
output consistent and lets the skill focus on gathering good data.

Input JSON shape (see references/linkedin-mechanics.md for the scraper that produces it):
{
  "searcher": "Name (run on their LinkedIn)",
  "value_prop": "One-line pitch used in the connector ask.",
  "results": [
    {"company": "Cleveland Kitchen", "people": [
      {"name": "Mac Anderson", "deg": "2nd", "headline": "Co-Founder ...",
       "mut": "Katie Dunn, Angelika O'Reilly and 7 other mutual connections"}
    ]}
  ]
}

Usage:
  python build_connector_sheet.py scrape.json "Company - Warm Intro Map.xlsx" \
      --value-prop "We help X do Y."
"""
import argparse, json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY="1F2A44"; ACCENT="2E7D5B"; LIGHT="EAF2EE"; GREY="6B7280"


def parse_mut(m):
    """Return (named_connectors, hidden_count) from a LinkedIn mutuals string."""
    if not m:
        return [], 0
    h = re.search(r"and (\d+) other mutual connection", m)
    hidden = int(h.group(1)) if h else 0
    s = re.sub(r"\s+and \d+ other mutual connections?", "", m)
    s = re.sub(r"\s+(is|are) (a )?mutual connections?", "", s)
    s = re.sub(r"\s+and ", ", ", s)
    names = [n.strip() for n in s.split(",") if n.strip()]
    return names, hidden


def style_header(ws, heads, widths):
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    fill = PatternFill("solid", fgColor=NAVY)
    thin = Side(style="thin", color="D9DCE1")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, (h, w) in enumerate(zip(heads, widths), 1):
        c = ws.cell(1, i, h)
        c.font = hf; c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="center"); c.border = bd
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    return bd


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input")
    ap.add_argument("output")
    ap.add_argument("--value-prop", default="")
    args = ap.parse_args()

    data = json.load(open(args.input))
    value_prop = args.value_prop or data.get("value_prop", "")
    searcher = data.get("searcher", "the user")

    base = Font(name="Arial", size=10)
    bold = Font(name="Arial", size=10, bold=True)
    sub = PatternFill("solid", fgColor=LIGHT)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="D9DCE1")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)

    targets, connector_map = [], {}
    for comp in data["results"]:
        for p in comp.get("people", []):
            named, hidden = parse_mut(p.get("mut", ""))
            targets.append({"company": comp["company"], "name": p["name"], "deg": p.get("deg", ""),
                            "headline": p.get("headline", ""), "named": named, "hidden": hidden,
                            "total": len(named) + hidden})
            for c in named:
                connector_map.setdefault(c, []).append(f'{p["name"]} ({comp["company"]})')

    wb = Workbook()

    ws = wb.active; ws.title = "Read me"; ws.column_dimensions["A"].width = 110
    found_brands = [c for c in data["results"] if c.get("people")]
    empty = [c["company"] for c in data["results"] if not c.get("people")]
    lines = [
        ("Warm intro map", True, 16, NAVY),
        ("", False, 10, None),
        ("What this is", True, 12, ACCENT),
        (f"A map of one network into a set of target companies. Run on {searcher}. For each target it shows who they are 1st or 2nd-degree connected to, and which of their contacts (connectors) can open that door.", False, 10, None),
        ("", False, 10, None),
        ("How to use it", True, 12, ACCENT),
        ("Start on 'By connector', sorted by reach. Pick a connector, send Email 1 (see the Email 1 tab) listing who they're connected to. Once they say yes, send the forwardable blurbs (Email 2, use the forwardable-intro skill).", False, 10, None),
        ("", False, 10, None),
        ("Pitch used in the ask", True, 12, ACCENT),
        (value_prop, False, 10, None),
        ("", False, 10, None),
        ("Two things to know", True, 12, ACCENT),
        ("1. Whose network: the connectors are whoever was logged in when this ran. A different person running it gets their own map.", False, 10, None),
        ("2. Hidden depth: LinkedIn names ~2 mutuals per target and hides the rest behind '+N others'. So named connectors are the tip; the 'Total warm paths' column shows where the real depth is. Expand a shortlist by clicking each target's '+N others' on LinkedIn.", False, 10, None),
        ("", False, 10, None),
        ("Coverage", True, 12, ACCENT),
        (f"{len(targets)} warm contacts across {len(found_brands)} of {len(data['results'])} targets." + (f" No warm path found for: {', '.join(empty)}." if empty else ""), False, 10, None),
    ]
    for r, (text, b, sz, color) in enumerate(lines, 1):
        c = ws.cell(r, 1, text)
        c.font = Font(name="Arial", bold=b, size=sz, color=(color or "000000"))
        c.alignment = wrap

    ws2 = wb.create_sheet("By connector")
    style_header(ws2, ["Connector", "Doors they can open", "Targets they're connected to"], [34, 18, 80])
    ranked = sorted(connector_map.items(), key=lambda kv: (-len(kv[1]), kv[0]))
    for r, (conn, tlist) in enumerate(ranked, 2):
        ws2.cell(r, 1, conn).font = bold
        ws2.cell(r, 2, len(tlist)).font = base
        ws2.cell(r, 3, "; ".join(sorted(set(tlist)))).font = base
        for col in range(1, 4):
            cell = ws2.cell(r, col); cell.alignment = wrap; cell.border = bd
            if r % 2 == 0: cell.fill = sub

    ws3 = wb.create_sheet("By target")
    style_header(ws3, ["Target", "Person", "Title", "Degree", "Named connectors", "Hidden others", "Total warm paths"], [20, 22, 46, 9, 40, 14, 16])
    for r, t in enumerate(sorted(targets, key=lambda x: (-x["total"], x["company"])), 2):
        vals = [t["company"], t["name"], t["headline"], t["deg"],
                ", ".join(t["named"]) or ("direct connection" if t["deg"] == "1st" else ""),
                t["hidden"], t["total"]]
        for col, v in enumerate(vals, 1):
            cell = ws3.cell(r, col, v); cell.font = base; cell.alignment = wrap; cell.border = bd
            if r % 2 == 0: cell.fill = sub

    ws5 = wb.create_sheet("Email 1"); ws5.column_dimensions["A"].width = 100
    email = [
        ("Email 1 — the connector ask", True, 14, NAVY),
        ("Send to a connector from 'By connector'. Swap in their name and the targets they know. Keep it short.", False, 10, GREY),
        ("", False, 10, None),
        ("Subject: quick intro favor", True, 11, None),
        ("", False, 10, None),
        ("hey [connector],", False, 10, None),
        ("", False, 10, None),
        ("saw you're connected to a few folks at companies i'm helping out: [targets].", False, 10, None),
        ("", False, 10, None),
        (value_prop.lower() if value_prop else "[one-line pitch].", False, 10, None),
        ("", False, 10, None),
        ("if i send you a couple short forwardable blurbs, would you mind passing them along? totally fine if some aren't the right fit.", False, 10, None),
        ("", False, 10, None),
        ("thanks so much,", False, 10, None),
        ("[name]", False, 10, None),
        ("", False, 10, None),
        ("Email 2 = the forwardable blurbs. Generate with the forwardable-intro skill once the connector says yes.", False, 10, ACCENT),
    ]
    for r, (text, b, sz, color) in enumerate(email, 1):
        c = ws5.cell(r, 1, text)
        c.font = Font(name="Arial", bold=b, size=sz, color=(color or "000000")); c.alignment = wrap

    wb.save(args.output)
    print(f"saved {args.output}: {len(connector_map)} connectors, {len(targets)} targets")
    if ranked:
        print("top connectors:", [(c, len(t)) for c, t in ranked[:8]])


if __name__ == "__main__":
    main()
